import json
import uuid

import openpyxl

sheet = None

destination_uuid = []

type_column_number = None
from_column_number = None
condition_column_number = None
save_name_column_number = None
text_column_number = None
media_column_number = None
condition_var_column_number = None
choices_column_numbers = []

choices = []

checked_condition_columns = []


def get_maximum_rows():
    global sheet
    return sheet.max_row


def get_maximum_columns():
    global sheet
    return sheet.max_column


def read_sheet_detail(sheet_name):
    global sheet
    compatible_file = openpyxl.load_workbook('C:\\Users\\Usman Ali\\Downloads\\ehmad_test_chat_flows.xlsx')
    sheet = compatible_file[sheet_name]


def get_sheet_cell_detail(row, column):
    global sheet
    return sheet.cell(row=row, column=column).value


def generate_uuid():
    return str(uuid.uuid4())


def get_last_node_detail(mark_as_completed, sheet_name):
    last_node_detail = {
        'uuid': generate_uuid(),
        'actions': [], 'exits': []
    }

    last_action_detail = {
        'uuid': generate_uuid(),
        'type': 'set_contact_field', 'field': {'key': f'{sheet_name}__completed',
                                               'name': f'{sheet_name}__completed'}, 'value': 'true'
    }

    last_exit_detail = {
        'uuid': generate_uuid(), 'destination_uuid': None
    }

    if mark_as_completed:
        last_action_detail['field']['key'] = 'task_relax__completed'
        last_action_detail['field']['name'] = 'task_relax__completed'
        last_exit_detail['destination_uuid'] = '0367ab86-a4a3-4116-88dd-10d36a17f829'

    last_node_detail['actions'].append(last_action_detail)
    last_node_detail['exits'].append(last_exit_detail)

    return last_node_detail


def get_required_column_numbers(sheet_name):
    global type_column_number, from_column_number, condition_column_number, text_column_number, \
        media_column_number, save_name_column_number, choices_column_numbers, condition_var_column_number

    choices_column_numbers = []
    condition_var_column_number = []

    read_sheet_detail(sheet_name)

    for column in range(1, get_maximum_columns()):
        first_row = get_sheet_cell_detail(1, column)

        if first_row == 'type':
            type_column_number = column
        elif first_row == 'from':
            from_column_number = column
        elif first_row == 'condition':
            condition_column_number = column
        elif first_row == 'condition_var':
            condition_var_column_number = column
        elif first_row == 'message_text':
            text_column_number = column
        elif first_row == 'media':
            media_column_number = column
        elif first_row == 'save_name':
            save_name_column_number = column
        elif first_row == 'choice_1' or first_row == 'choice_2' or first_row == 'choice_3':
            choices_column_numbers.append(column)
        elif not first_row:
            break


def get_condition_node_detail(row, condition_values, save_name):
    global save_name_column_number, condition_var_column_number
    global destination_uuid

    condition_node_detail = {
        'uuid': destination_uuid[-1] if destination_uuid else generate_uuid(),
        'actions': [], 'router': {'type': 'switch', 'cases': [],
                                  'categories': [{'exit_uuid': generate_uuid(),
                                                  'name': 'All Responses',
                                                  'uuid': generate_uuid()}],
                                  'operand': '@input.text',
                                  'default_category_uuid': '',
                                  'wait': {'type': 'msg'}},
        'exits': [{'uuid': 'e59fda75-21e1-4d77-83a7-7733cc721eff', 'destination_uuid': None}],
    }

    if destination_uuid:
        destination_uuid.remove(destination_uuid[-1])

    condition_node_detail['router']['default_category_uuid'] = condition_node_detail['router']['categories'][0]['uuid']
    condition_node_detail['exits'][0]['uuid'] = condition_node_detail['router']['categories'][0]['exit_uuid']

    if save_name:
        condition_node_detail['router']['result_name'] = save_name
    else:
        for condition_text in condition_values:
            cases_detail = {
                'arguments': [], 'category_uuid': generate_uuid(),
                'type': 'has_only_phrase', 'uuid': generate_uuid()
            }

            categories_detail = {
                'exit_uuid': generate_uuid(), 'name': '',
                'uuid': cases_detail['category_uuid']
            }

            exits_detail = {
                'uuid': categories_detail['exit_uuid'],
                'destination_uuid': generate_uuid()
            }

            destination_uuid.insert(0, exits_detail['destination_uuid'])

            if condition_text == 'Unticked Value' or condition_text == 'No':
                case_detail = {
                    'arguments': [], 'category_uuid': generate_uuid(),
                    'type': 'has_only_phrase', 'uuid': generate_uuid()
                }

                categorie_detail = {
                    'exit_uuid': generate_uuid(),
                    'name': '', 'uuid': generate_uuid()
                }

                exit_detail = {
                    'uuid': 'e59fda75-21e1-4d77-83a7-7733cc721eff',
                    'destination_uuid': '65355f1a-f702-48c3-a015-1774112607e5',
                }

                case_detail['arguments'].append(str(condition_text))
                categorie_detail['name'] = str(condition_text)

                condition_node_detail['router']['cases'].append(case_detail)
                condition_node_detail['router']['categories'].append(categorie_detail)
                condition_node_detail['exits'].append(exit_detail)

            if condition_text:
                categories_detail['name'] = str(condition_text)
                cases_detail['arguments'].append(str(condition_text))

                condition_node_detail['router']['cases'].append(cases_detail)
                condition_node_detail['router']['categories'].append(categories_detail)
                condition_node_detail['exits'].append(exits_detail)

        if save_name_column_number:
            if get_sheet_cell_detail(row, save_name_column_number):
                condition_node_detail['router'].pop('wait', None)
                condition_node_detail['router'][
                    'operand'] = f'@fields.{get_sheet_cell_detail(row, save_name_column_number)}'

        if condition_var_column_number:
            if get_sheet_cell_detail(row=row, column=condition_var_column_number):
                condition_node_detail['router'].pop('wait', None)
                condition_node_detail['router']['operand'] = get_sheet_cell_detail(row=row,
                                                                                   column=condition_var_column_number)

    return condition_node_detail


def get_condition_values(row):
    global condition_column_number, from_column_number, type_column_number
    global checked_condition_columns

    condition_column_values = [get_sheet_cell_detail(row, condition_column_number)]

    from_column_value = get_sheet_cell_detail(row, from_column_number)
    increment_row = 1

    while True:
        if from_column_value == get_sheet_cell_detail(row + increment_row, from_column_number):
            checked_condition_columns.append(row + increment_row)
            condition_column_values.append(get_sheet_cell_detail(row + increment_row, condition_column_number))
            increment_row = increment_row + 1

        elif get_sheet_cell_detail(row + increment_row, type_column_number) == 'mark_as_completed':
            increment_row = increment_row + 1
        else:
            break

    return condition_column_values


def get_message_text_node_detail(row):
    global choices_column_numbers, media_column_number
    global choices
    global destination_uuid

    choices = []

    message_text_node_detail = {
        'uuid': destination_uuid[-1] if destination_uuid else generate_uuid(),
        'actions': [], 'exits': [],
    }

    if destination_uuid:
        destination_uuid.remove(destination_uuid[-1])

    message_text_action_detail = {
        'attachments': [],
        'text': get_sheet_cell_detail(row, text_column_number),
        'type': 'send_msg', 'quick_replies': [],
        'uuid': generate_uuid()
    }

    if get_sheet_cell_detail(row=row, column=media_column_number):
        message_text_action_detail['attachments'].append('image:' + get_sheet_cell_detail(row=row, column=media_column_number))

    message_text_exist_detail = {
        'uuid': generate_uuid(),
        'destination_uuid': generate_uuid()
    }

    if get_sheet_cell_detail(row + 1, text_column_number) == 2:
        message_text_exist_detail['destination_uuid'] = None

    destination_uuid.insert(0, message_text_exist_detail['destination_uuid'])

    for choice in choices_column_numbers:
        choice_text = get_sheet_cell_detail(row, choice)

        if choice_text:
            choices.append(choice_text)

    for choice_text in choices:
        message_text_action_detail['quick_replies'].append(choice_text)

    message_text_node_detail['actions'].append(message_text_action_detail)
    message_text_node_detail['exits'].append(message_text_exist_detail)

    return message_text_node_detail


def get_save_name_node_detail(save_name_column_value, row):
    global type_column_number, text_column_number

    save_name_node_detail = {
        'uuid': '46084334-4111-4866-897a-b189a3e2d51c', 'actions': [{'uuid': '8df481e6-c0a5-4ee5-916b-597163597039',
                                                                     'type': 'set_contact_field',
                                                                     'field': {'key': f'{save_name_column_value}',
                                                                               'name': f'{save_name_column_value}'},
                                                                     'value': f'@results.{save_name_column_value}'}],
        'exits': [{'uuid': 'ee6809e2-1336-49d0-913f-5d18ffd484da',
                   'destination_uuid': '1ea89c1e-aebd-4267-9fff-6d419a101d47'}],
    }

    if get_sheet_cell_detail(row, type_column_number) == 'save_value':
        save_name_node_detail['actions'][0]['value'] = get_sheet_cell_detail(row, text_column_number)

    return save_name_node_detail


def get_all_nodes_detail(flows_detail, sheet_name):
    global sheet
    global type_column_number, from_column_number, condition_column_number, \
        text_column_number, save_name_column_number, choices_column_numbers
    global checked_condition_columns
    global choices

    checked_condition_columns = []
    get_required_column_numbers(sheet_name)

    for row in range(2, get_maximum_rows() + 1):
        if not get_sheet_cell_detail(row, 2):
            break

        if get_sheet_cell_detail(row, type_column_number) == 'mark_as_completed':
            flows_detail['nodes'].append(get_last_node_detail(True, sheet_name))
        else:
            if get_sheet_cell_detail(row, condition_column_number) and row not in checked_condition_columns:
                checked_condition_columns.append(row)
                flows_detail['nodes'].append(get_condition_node_detail(row, get_condition_values(row), None))

            if get_sheet_cell_detail(row, text_column_number) and get_sheet_cell_detail(row, text_column_number) != 2:
                if not get_sheet_cell_detail(row, type_column_number) == 'save_value':
                    flows_detail['nodes'].append(get_message_text_node_detail(row))

                if choices and get_sheet_cell_detail(row=row + 1, column=2) is None:
                    flows_detail['nodes'].append(get_condition_node_detail(row, choices, None))

            if save_name_column_number:
                if get_sheet_cell_detail(row, save_name_column_number):
                    save_name_column_value = get_sheet_cell_detail(row, save_name_column_number)

                    if not get_sheet_cell_detail(row, type_column_number) == 'save_value':
                        flows_detail['nodes'].append(get_condition_node_detail(row, [], save_name_column_value))
                        flows_detail['nodes'].append(get_save_name_node_detail(save_name_column_value, row))
                    else:
                        flows_detail['nodes'].append(get_save_name_node_detail(save_name_column_value, row))


def get_detail_in_flows(sheet_name):
    flows_detail = {
        'name': f'{sheet_name}', 'uuid': generate_uuid(),
        'spec_version': '13.1.0', 'language': 'base', 'type': 'messaging',
        'nodes': [], '_ui': None, 'revision': 0, 'expire_after_minutes': 60,
        'metadata': {'revision': 0}, 'localization': {}
    }

    if get_all_nodes_detail(flows_detail, sheet_name):
        flows_detail['nodes'].append(get_all_nodes_detail(flows_detail, sheet_name))

    # flows_detail['nodes'].append(get_last_node_detail(False, sheet_name))

    return flows_detail


if __name__ == '__main__':
    # sheets = ['example_habit', 'example_characters', 'example_mark_as_completed',
    #           'example_story2', 'example_story1', 'example_user_input', 'example_variables',
    #           'example_tickbox_2', 'example_tickbox', 'example_long_xxxxxxxxxxxxxxxx']

    sheets = ['example_story1']

    for sheet_name in sheets:
        complete_sheet_detail = {
            'campaigns': [], 'fields': [], 'flows': [], 'groups': [],
            'site': 'https://rapidpro.idems.international',
            'triggers': [], 'version': '13',
        }

        complete_sheet_detail['flows'].append(get_detail_in_flows(sheet_name))

        with open(f'{sheet_name}.json', 'w') as sheet_detail:
            json.dump(complete_sheet_detail, sheet_detail)
